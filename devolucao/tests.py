from django.test import TestCase
from openpyxl import load_workbook
from devolucao.models import Livro

# # Create your tests here.
# wb = load_workbook('DEVOLREC.xlsx')
# # for i in range(1,2415):
# for sheet in wb:
#     for i in range(2,2416):
#         livro = Livro()
#         for j in range(1,5):
#             if j==1:
#                 livro.isbn = sheet.cell(row=i, column=j).value
#             elif j==2:
#                 # print(sheet.cell(row=i, column=j).value)
#                 livro.quantidade_devolver = int(sheet.cell(row=i, column=j).value)
#             elif j==3:
#                 # print(sheet.cell(row=i, column=j).value)
#                 livro.quantidade_atual = int(sheet.cell(row=i, column=j).value)
#             elif j==4:
#                 livro.titulo = sheet.cell(row=i, column=j).value
#         livro.save()

# livros = Livro.objects.all()
# for livro in livros:
#     if len(Livro.objects.all().filter(isbn=livro.isbn))>1:
#         print(livro.id)